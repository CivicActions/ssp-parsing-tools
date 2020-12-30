#!/usr/bin/env python3

# TODO: We need to make this more functional. Currently the spreadsheets are
# hardcoded and the queries are based on JOINs specific to the spreadsheets.

import json
from pathlib import Path

import dataset
from openpyxl import load_workbook

db = dataset.connect('sqlite:///data.db')

def main():
    sp = {
        'inventory': 'inventory.xlsx',
        'tracking': 'tracking.xlsx',
        'implements': 'implementation.xlsx'
    }
    for t, f in sp.items():
        read_sheet(t, f)

    create_projects()

def read_sheet(t, f):
    # Read the spreadsheets and write the rows into the Sqlite3 DB.
    print('Writing {} to database table {}.'.format(f, t))
    db.begin()
    wb = load_workbook(f, read_only=True)
    sheet = wb.active
    maxColumns = sheet.max_column
    maxRows = sheet.max_row
    db.create_table(t)

    # Create keys from column headers and make them more SQL friendly.
    headers = [cell.value.replace(" ", "_").lower() for cell in next(sheet.rows)]
    rdict = {}
    for row in sheet.iter_rows(min_row=2, max_row=maxRows, min_col=1, max_col=maxColumns, values_only=True):
        for x, y in enumerate(row):
            rdict[headers[x]] = y
        try:
            db[t].insert(rdict)
            db.commit()
        except:
            db.rollback()

def create_projects():
    # Read the data from the database for each unique project and write it
    # to a JSON file.
    projects = db['tracking'].distinct('authorization_package_name')
    for p in projects:
        project = {}
        project['name'] = p['authorization_package_name']
        project['acronym'] = get_project_acronym(p['authorization_package_name'])
        project['meta'] = get_project_metadata(p['authorization_package_name'])
        project['controls'] = get_project_controls(p['authorization_package_name'])
        write_json(project)

def get_project_acronym(name):
    result = db['implements'].find_one(information_system_or_program_name=name)
    try:
        result['acronym']
        return result['acronym']
    except NameError:
        print("No acronym for project {}".format(name))


def get_project_metadata(name):
    result = db['inventory'].find_one(information_system_or_program_name=name)
    return result

def get_project_controls(name):
    query = str('''SELECT
            t.control_number,
            t.control_name,
            t.control_set_version_number,
            t.allocation_status,
            t.overall_update_date,
            t.overall_control_status,
            t.assessment_status,
            t.tracking_id,
            i.shared_implementation_details,
            i.private_implementation_details,
            i.tracking_id AS tracking
        FROM
            tracking t
        JOIN
            implements i
        ON
            t.authorization_package_name = i.information_system_or_program_name
        AND
            t.control_number = i.control_number
        WHERE
            t.authorization_package_name = "{}"'''.format(name))
    controls = {}
    result = db.query(query)
    for row in result:
        controls[row['control_number']] = row
    return controls

def write_json(project):
    out_dir = Path("projectjson")
    out_dir.mkdir(exist_ok=True)
    fname = project['acronym'].replace('/', '').replace(' ', '')
    print('Creating JSON file {}'.format(fname + '.json'))
    json_project = json.dumps(project, indent=4)
    j = out_dir.joinpath(fname + '.json')
    j.touch()
    j.write_text(json_project)


if __name__ == "__main__":
    # Remove the database file so that we don't duplicate records.
    d = Path('data.db')
    d.unlink(missing_ok=True)
    main()
